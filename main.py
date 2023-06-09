import openpyxl
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
import os
from PyPDF2 import PdfFileMerger, PdfMerger
from reportlab.lib.pagesizes import portrait, A4
from reportlab.pdfgen.canvas import Canvas

# 注册 TrueType 字体
pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))

# 设置页码和文件名初始值
page_num = 1
filename = f'output/example{page_num}.pdf'

# 创建 Canvas 对象
canvas = Canvas(filename, pagesize=portrait(A4))

# 设置字体
canvas.setFont('SimSun', 49)

# 读取 SKU 和箱数信息
wb = openpyxl.load_workbook('prepare.xlsx')
ws = wb.active
sku_list = [cell.value for cell in ws['A']]
box_num_list = [cell.value for cell in ws['B']]
total_box_num = sum(box_num_list)

# 循环向 A1 单元格中写入数据
current_box = 1
for i, sku in enumerate(sku_list):
    for j in range(box_num_list[i]):
        # 构造字符串并填入 A1 单元格
        sku_str = f'{sku}'
        content = f'{sku_str}'

        # 获取页面大小
        width, height = portrait(A4)

        # 计算文本框的位置
        text_width, text_height = canvas.stringWidth(content, 'SimSun', 49), canvas._leading * len(content.split('\n'))
        x = (width - text_width) / 2
        y = (height - text_height) / 2

        # 绘制文本框
        textobject = canvas.beginText(x, y)
        textobject.setFont('SimSun', 49)
        textobject.textLines(content)
        canvas.drawText(textobject)

        # 保存 PDF 文件
        canvas.save()

        # 更新页码和文件名和 Canvas 对象
        page_num += 1
        filename = f'output/example{page_num}.pdf'
        canvas = Canvas(filename, pagesize=portrait(A4))
        canvas.setFont('SimSun', 49)

        current_box += 1
